import pandas as pd
import re
import time
import yaml
import logging
import testlink
import vodafone

from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from datetime import datetime
from docx import Document

config = yaml.load(open('Testlink_config.yml'), Loader=yaml.Loader)
time_string = datetime.now().strftime('%Y-%m-%d')
dev_logger: logging.Logger = logging.getLogger(name='dev')
dev_logger.setLevel(logging.DEBUG)

formatter: logging.Formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s', datefmt='%Y%m%d %H:%M:%S')

# print on console
handler: logging.StreamHandler = logging.StreamHandler()    #StreamHandler 用來控制輸出終端機的相關設定
handler.setFormatter(formatter)
dev_logger.addHandler(handler)

#save on log
file_handler = logging.StreamHandler(open(f'{time_string}.log', 'w'))
file_handler.setLevel(logging.WARNING)
file_handler.setFormatter(formatter)
dev_logger.addHandler(file_handler)

GetRequirement_data={}
NameFromePDF_data={}

WordToExcel = config['WordToExcel']
Excel = config['Excel']
GetRequirement_data_key = []
GetRequirement_data_value = []
GetRequirement_data_name = []
NewGetRequirement=[]
not_found = []
end_of_excel = f'{time_string}_mapping_{Excel[0]}'

timeout = 20

border = Border(left=Side(border_style='thin',color='000000'),
right=Side(border_style='thin',color='000000'),
top=Side(border_style='thin',color='000000'),
bottom=Side(border_style='thin',color='000000'))

def main():
    start = time.time()
    # word_to_excel()
    # merge()
    # vodafone.find_from_vodafone()
    get_name()
    get_from_testlink()
    mapping()
    
    end = time.time()
    print('Time elapsed: ' + str(start-end) + ' seconds')

def get_from_testlink():        
    GetRequirement_data['Testlink']=[]
    
    # print(GetRequirement_data['Name'])
    Document_ID_text = testlink.find_from_testlink(GetRequirement_data['Name'])
    GetRequirement_data['Testlink'] = Document_ID_text
            
    # for key, value in GetRequirement_data.items():
    #     print(key, len([item for item in value if item]))
    
    Testlink_data = pd.DataFrame(GetRequirement_data)
    Testlink_data.to_excel(end_of_excel, index=False)

def mapping():
    mapping = load_workbook(end_of_excel, read_only = False)
    mapping_sheet = mapping[mapping.sheetnames[0]]
    string_list = [str(element) for element in sorted(not_found)]
    result_string = ','.join(string_list)
    mapping_sheet.cell(row=1, column=mapping_sheet.max_column+1, value = 'Not_found').border  = border
    mapping_sheet.cell(row=1, column=mapping_sheet.max_column, value = 'Not_found').font  = Font(bold=True)
    mapping_sheet.cell(row=2, column=mapping_sheet.max_column, value = len(string_list)).font = Font(bold=True, color="e06666", size=15)
    mapping_sheet.cell(row=2, column=mapping_sheet.max_column+1, value = result_string)
    
    #將excel 所有儲存格 對齊 水平對齊 填滿
    for ws in mapping.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                align = Alignment(horizontal='fill')
                cell.alignment = align
                
    mapping.close()
    mapping.save(end_of_excel)
    
def word_to_excel():
    for e in range(2):  
        word = config['Word'][e]
        doc = Document(word) 
        xls = pd.ExcelWriter(config['WordToExcel'][e]) 
        
        tables = doc.tables  # word 中所有 tables
        
        for i, tb in enumerate(tables):
            mat = []  
            for r in range(0, len(tb.rows)):
                row = []
                for c in range(0, len(tb.columns)):
                    cell = tb.cell(r, c)
                    txt = cell.text if cell.text != '' else ' '  # 無內容用空格佔位
                    row.append(txt)
                mat.append(row)
            
            df = pd.DataFrame(mat)
            
            df.rename(columns={0:'Req Parent', 1:'Name', 2:'Description'}, inplace = True)
            df.to_excel(xls, sheet_name=f'{i}', index=False, header=True)      # 並且不要首列 index 和首行的 header（預設會有，例如數字標號）

        xls.close()  
        
        # 多sheet合併
        dfs = pd.read_excel(WordToExcel[e], sheet_name=None)
        df = pd.concat(dfs.values(), ignore_index=True)
        sheet = WordToExcel[e].replace('.xlsx','')[-9:]
        df.to_excel(f'{WordToExcel[e]}', sheet_name=sheet, index=False, header=True)
        
def merge():
    df_a = pd.read_excel(WordToExcel[0])
    df_b = pd.read_excel(WordToExcel[1])

    merged_df = pd.concat([df_a, df_b], ignore_index=True)
    sheet = WordToExcel[0].replace('.xlsx','')[-8:-3]+WordToExcel[1].replace('.xlsx','')[-3:]
    merged_df.to_excel('merged.xlsx', sheet_name=sheet,index=False)
    
def get_name():
    wb1 = load_workbook(Excel[0], read_only = False)
    wb2 = load_workbook(Excel[1], read_only = False)
    
    for xlsx in range(2):      
        if xlsx == 0:
            get_excel =  wb1
        else:
            get_excel =  wb2
            
        for sheet in range(len(get_excel.sheetnames)):
            work1 = wb1[wb1.sheetnames[sheet]]
            work2 = wb2[wb2.sheetnames[sheet]]
            
        make_dict(work1, work2)
    
    k=0
    count_list=[]
    GetRequirement_data['Name']=[]    
    for key in GetRequirement_data:
        if key == 'Description':
            for i in range(1, work1.max_row):
                if len(count_list) == 0:
                    GetRequirement_data['Name'].append('')
                    not_found.append(work1.cell(i+1 , k).row)
                    # print(f'第 {i} 列，加入空白，list數{len(GetRequirement_data['Name'])}')
                    
                count_list.clear()
                for b, description in enumerate(NameFromePDF_data.get(key)):
                    if is_blank_or_none(description) == False:
                        # print(work1.cell(i+1 , k).row)
                        test_str = re.sub(r'[^a-zA-Z0-9]', ' ', GetRequirement_data[key][i-1][:len(GetRequirement_data[key][i-1])//4])
                        if re.search(test_str, description):
                            print(f'查看 {Excel[1]} 中 第 {b+2} 列有相似的 {GetRequirement_data[key][i-1]}') #第一個excel 從1開始，第二個excel從0開始，整體少2，故+2   
                            # print(work2.cell(b+2, k).value)  
                            if 'KIP-REQ' in work2.cell(b+2, k).value:
                                GetRequirement_data['Name'].append(work2.cell(b+2, k).value)

                                count_list.append(b+2)
                                # print(len(count_list))
                                if len(count_list) != 1:
                                    GetRequirement_data[list(GetRequirement_data.keys())[0]].insert(i, '')     #Req Parent
                                    GetRequirement_data[list(GetRequirement_data.keys())[1]].insert(i, '')      #Summary
                                    GetRequirement_data[list(GetRequirement_data.keys())[2]].insert(i, '')      #Description

                                # print(f'第 {i} 列，加入 {len(count_list)}，list數{len(GetRequirement_data['Name'])}')

        k+=1
        
def is_blank_or_none(s):
    return s is None or len(s.strip()) == 0

def make_dict(work1, work2):
    for r in range(1, work1.max_column+1):
        key = work1.cell(1, r).value
        GetRequirement_data[key] = []
        for c in range(2, work1.max_row+1):
            value = work1.cell(c, r).value
            GetRequirement_data[key].append(value)
    # print(GetRequirement_data)
    
    for r in range(1, work2.max_column+1):
        key = work2.cell(1, r).value
        NameFromePDF_data[key] = []
        for c in range(2, work2.max_row+1):
            value = work2.cell(c, r).value
            NameFromePDF_data[key].append(value)
    # print(NameFromePDF_data)

if __name__ == "__main__":
    main()