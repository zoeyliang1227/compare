import pandas as pd
import time
import openpyxl
import warnings
import re
import yaml

from openpyxl import load_workbook
from datetime import datetime
from docx import Document

config = yaml.load(open('Testlink_config.yml'), Loader=yaml.Loader)
time_string = datetime.now().strftime('%Y-%m-%d')

dict1={}
dict2={}

testlink = []

def main():
    start = time.time()
    # word_to_excel()
    # merge()
    
    get_name()
    
    
    
    end = time.time()
    print('Time elapsed: ' + str(start-end) + ' seconds')
    
def word_to_excel():
    for e in range(2):  
        word = config['get_word'][e]
        doc = Document(word) 
        xls = pd.ExcelWriter(config['get_excel'][e]) 
        
        tables = doc.tables  # word 中所有 tables
        
        for i, tb in enumerate(tables):
            mat = []  
            for r in range(0, len(tb.rows)):
                row = []
                for c in range(0, len(tb.columns)):
                    cell = tb.cell(r, c)
                    txt = cell.text if cell.text != '' else ' '  # 無內容用空格佔位
                    row.append(txt)
                mat.append(row)
            
            df = pd.DataFrame(mat)
            
            df.rename(columns={0:'Req Parent', 1:'Name', 2:'Description'}, inplace = True)
            df.to_excel(xls, sheet_name=f'{i}', index=False, header=True)      # 並且不要首列 index 和首行的 header（預設會有，例如數字標號）

        xls.close()  
        
        # 多sheet合併
        dfs = pd.read_excel(config['get_excel'][e], sheet_name=None)
        df = pd.concat(dfs.values(), ignore_index=True)
        sheet = config['get_excel'][e].replace('.xlsx','')[-9:]
        df.to_excel(f'{config['get_excel'][e]}', sheet_name=sheet, index=False, header=True)
        
def merge():
    df_a = pd.read_excel(config['get_excel'][0])
    df_b = pd.read_excel(config['get_excel'][1])

    merged_df = pd.concat([df_a, df_b], ignore_index=True)
    sheet = config['get_excel'][0].replace('.xlsx','')[-8:-3]+config['get_excel'][1].replace('.xlsx','')[-3:]
    merged_df.to_excel('merged.xlsx', sheet_name=sheet,index=False)

def make_dict(work1, work2):
    for r in range(1, work1.max_column+1):
        key = work1.cell(1, r).value
        dict1[key] = []
        for c in range(2, work1.max_row+1):
            value = work1.cell(c, r).value
            dict1[key].append(value)
    # print(dict1)
    
    for r in range(1, work2.max_column+1):
        key = work2.cell(1, r).value
        dict2[key] = []
        for c in range(2, work2.max_row+1):
            value = work2.cell(c, r).value
            dict2[key].append(value)
    # print(dict2)
    
def get_name():
    wb1 = load_workbook(config['excel'][0], read_only = False)
    wb2 = load_workbook(config['excel'][1], read_only = False)
    
    for xlsx in range(2):      
        if xlsx == 0:
            get_excel =  wb1
            wb_excel = config['excel'][0]
        else:
            get_excel =  wb2
            wb_excel = config['excel'][1]
            
        for sheet in range(len(get_excel.sheetnames)):
            work1 = wb1[wb1.sheetnames[sheet]]
            work2 = wb2[wb2.sheetnames[sheet]]
            
        make_dict(work1, work2)
    
    k=0
    for key in dict1:
        if key == 'Description':
            for i in range(1, work1.max_row):
                for b, bb in enumerate(dict2.get(key)):
                    if  bb is not None:
                        a = re.search(rf'{dict1[key][i-1][:23]}', bb)
                        if a != None:
                            print(f'{dict1[key][i-1]}在 {config['excel'][1]} 中 第 {b+2} 列有相似的') #第一個excel 從1開始，第二個excel從0開始，整體少2，故+2    
                            # print(work2.cell(b+2, k).value)
                            testlink.append(work2.cell(b+2, k).value)
        k+=1

if __name__ == "__main__":
    main()