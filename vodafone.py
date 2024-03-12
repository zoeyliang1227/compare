import time
import yaml
import openpyxl
import re
import pandas as pd

from openpyxl import load_workbook
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common import exceptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys

config = yaml.load(open('Testlink_config.yml'), Loader=yaml.Loader)
timeout = 20
time_string = datetime.now().strftime('%Y-%m-%d')

Excel = config['Excel']

GetDescription_data={}
Req_parent=[]

def get_driver():
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')                 # 瀏覽器不提供可視化頁面
    options.add_argument('-no-sandbox')               # 以最高權限運行
    options.add_argument('--start-maximized')        # 縮放縮放（全屏窗口）設置元素比較準確
    options.add_argument('--disable-gpu')            # 谷歌文檔說明需要加上這個屬性來規避bug
    options.add_argument('--window-size=1920,1080')  # 設置瀏覽器按鈕（窗口大小）
    options.add_argument('--incognito')               # 啟動無痕

    driver = webdriver.Chrome(options=options)
    url = config['DG4278_url']

    # driver.implicitly_wait(10)
    # driver.get(url)
    # driver.delete_all_cookies() #清cookie
    
    # with open("cookies.yml", "r") as f:
    #     cookies = yaml.safe_load(f)
    #     for c in cookies:
    #         if 'domain' in c:
    #             c['domain'] = 'xxx'
    #         print(c)
    #         driver.add_cookie(c)

    driver.get(url)
    
    return driver

def find_from_vodafone():
    driver = get_driver()
    login(driver)
    get_name()
    GetDescription_data['Description']=[]
    for req in Req_parent:
        # print(req)
        WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, 'quickSearchInput'))).send_keys(req) 
        WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, 'quickSearchInput'))).send_keys(Keys.ENTER)
        
        print(f'{req} get description from vodafone')
        #Description
        if check_description(driver) == True:
            description_text = WebDriverWait(driver, timeout).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="descriptionmodule"]/div[2]'))).text.strip()
            GetDescription_data['Description'].append(description_text)
        else:
            GetDescription_data['Description'].append('')
            
    # print(GetDescription_data)
    print('Write to excel')
    Description_data = pd.DataFrame(GetDescription_data)
    Description_data.to_excel(Excel[0], index=False)

def login(driver):
    print('Waiting for login...')
    WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, 'i0116'))).send_keys(config['DG4278_username']) 
    WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, 'idSIButton9'))).click() 
    time.sleep(3)
    WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, 'i0118'))).send_keys(config['DG4278_password']) 
    WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, 'idSIButton9'))).click() 
    number = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, 'idRichContext_DisplaySign'))).text
    print(f'Please enter number on your phone: {number}')
    time.sleep(8)
    WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, 'idSIButton9'))).click() 
    print('Waiting for the website to load...')
    # cookie2 = driver.get_cookies() #取得登入後cookie
    # with open("cookies.yml", "w") as f:
    #     yaml.safe_dump(data=cookie2, stream=f)
    
def get_name():            
    wb4 = load_workbook(Excel[0], read_only = False)
    work4 = wb4[wb4.sheetnames[0]]
    make_dict(work4)
    
    for key in GetDescription_data:
        # print(key)
        if key == 'Req Parent':
            for i in range(1, work4.max_row):
                # print(GetDescription_data[key][i-1])
                Req_parent.append(GetDescription_data[key][i-1])


def make_dict(work1):
    for r in range(1, work1.max_column+1):
        key = work1.cell(1, r).value
        GetDescription_data[key] = []
        for c in range(2, work1.max_row+1):
            value = work1.cell(c, r).value
            GetDescription_data[key].append(value)
    # print(GetDescription_data)

def check_description(driver):
    try:
        driver.find_element(By.XPATH, '//*[@id="descriptionmodule-label"]')
        return True
    except:
        return False
    